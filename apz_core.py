import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from typing import List, Optional
from apz_pdf_parser import extract_apz_data


def build_dataframe(pdf_paths: List[str]) -> pd.DataFrame:
    data = extract_apz_data(pdf_paths)
    df = pd.DataFrame(data)
    return df


def export_to_excel(
    df: pd.DataFrame, 
    filename: str = "APZ_Zusammenfassung_2025.xlsx",
    columns: Optional[List[str]] = None
) -> None:
    """
    Exportiert das DataFrame als formatierte Excel-Datei.
    columns = Liste der Spalten, die exportiert werden sollen (aus den Checkboxen).
    Wenn None oder leer → alle Spalten werden exportiert.
    """
    # Nur ausgewählte Spalten behalten (falls vorhanden)
    if columns:
        existing = [col for col in columns if col in df.columns]
        if existing:
            df = df[existing]

    wb = Workbook()
    ws = wb.active
    ws.title = "APZ Daten"

    # Header fett schreiben
    header = list(df.columns)
    for col_num, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num, value=value)
        cell.font = Font(bold=True)

    # Datenzeilen schreiben
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 2, column=c_idx, value=value)

    wb.save(filename)