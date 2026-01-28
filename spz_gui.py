import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Daten aus PDFs extrahieren (hardcoded aus deinen Blättern; erweitere mit PDF-Parsing-Funktion unten)
data = [
    {'WE-Nr.': 'WE250378', 'Hersteller': 'Dillinger', 'Zeugnisnr.': '771473-006', 'Schmelze': '550003', 'Walztafel': '36892-01', 'Material': 'S460M+Z35', 'Art': 'Blech', 't (mm)': 20, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'TM, Ist-Masse 8725 kg, RT-Zug ok (REH 68–73%)'},
    {'WE-Nr.': 'WE250378', 'Hersteller': 'Dillinger', 'Zeugnisnr.': '771473-006', 'Schmelze': '550003', 'Walztafel': '36859', 'Material': 'S460M+Z35', 'Art': 'Blech', 't (mm)': 20, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'TM Pos. 01, RT-Zug ok (REH 68–73%)'},
    {'WE-Nr.': 'WE250354', 'Hersteller': 'Dillinger', 'Zeugnisnr.': '771473-004', 'Schmelze': '550002', 'Walztafel': '36735-01', 'Material': 'S460M+Z35', 'Art': 'Blech', 't (mm)': 18, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'TM, Masse 10360 kg, RT-Zug ok (71–74%)'},
    {'WE-Nr.': 'WE250354', 'Hersteller': 'Dillinger', 'Zeugnisnr.': '771473-004', 'Schmelze': '550002', 'Walztafel': '36739', 'Material': 'S460M+Z35', 'Art': 'Blech', 't (mm)': 18, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'TM Pos. 05, RT-Zug ok (61–74%)'},
    {'WE-Nr.': 'WE24K013', 'Hersteller': 'Voestalpine', 'Zeugnisnr.': 'A0375183', 'Schmelze': '858354', 'Walztafel': '145440/1', 'Material': 'S355J2+N Z35', 'Art': 'Blech', 't (mm)': 70, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'A3', 'Oberfläche': 'A3', 'Grenzabmaße': 'B', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'Normalized rolled, Masse 8734 kg'},
    {'WE-Nr.': 'WE241258', 'Hersteller': 'Ilsenburger', 'Zeugnisnr.': '1432778', 'Schmelze': '05076', 'Walztafel': '596930 3', 'Material': 'S460M', 'Art': 'Blech', 't (mm)': 8, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A UG3', 'Grenzabmaße': 'C', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'entf.', 'Bemerkung': 'TM, 21 Platten, Masse 18955 kg'},
    {'WE-Nr.': 'WE210683', 'Hersteller': 'Voestalpine', 'Zeugnisnr.': 'A0369891', 'Schmelze': '839467', 'Walztafel': '286065/1', 'Material': 'S460M Z35', 'Art': 'Blech', 't (mm)': 30, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'Fine grained, Masse 4495 kg'},
    {'WE-Nr.': 'WE210683', 'Hersteller': 'Voestalpine', 'Zeugnisnr.': 'A0369891', 'Schmelze': '841158', 'Walztafel': '287334/1', 'Material': 'S460M Z35', 'Art': 'Blech', 't (mm)': 16, 'APZ': '3.2', 'DBS': 'ok', 'Schmelzanalyse': 'ok', 'Mech. Kennw.': 'ok', 'UT E1/S1 >10mm': 'S1', 'Oberfläche': 'A3', 'Grenzabmaße': 'A', 'Aubitz t>30': 'entf.', 'Radioaktiv.': 'frei', 'CE': 'ok', 'C-Güte THS': 'ok', 'Z35': 'ok', 'Bemerkung': 'Fine grained, Masse 3726 kg'}
]

df = pd.DataFrame(data)

# Excel erstellen
wb = Workbook()
ws = wb.active
ws.title = "APZ Daten"

# Header hinzufügen und fett formatieren
header = list(df.columns)
for col_num, value in enumerate(header, 1):
    cell = ws.cell(row=1, column=col_num, value=value)
    cell.font = Font(bold=True)

# Daten hinzufügen
for r_idx, row in df.iterrows():
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx + 2, column=c_idx, value=value)

self.update_button = QPushButton("Update")
self.update_button.clicked.connect(self.update_app)
layout.addWidget(self.update_button)
def update_app(self):
    subprocess.check_call(["git", "pull"])  # Wenn Git-repo
    self.log_text.append("Update abgeschlossen.")


wb.save("APZ_Zusammenfassung_2025.xlsx")
print("Datei erstellt.")  # Test-Ausgabe